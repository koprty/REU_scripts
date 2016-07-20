import openpyxl

days = ["Sun ","Mon ","Tue ","Wed ","Thu ","Fri ","Sat "]
first_tweet_count = 0.
total = 0.

wb = openpyxl.load_workbook("tweets_similar_to_retweeted.xlsx")
ws = wb.active
'''
wb.create_sheet("new_sheet",index=0)
new_sheet = wb.get_sheet_by_name("new_sheet")

r = 1
for row in old_ws.rows:
	if row[0].value=="ORIGINAL TWEET":
		new_sheet.cell(row = r, column = 1).value = ''
		r += 1
	for i in range(1,max_col+1):
		new_sheet.cell(row=r,column=i).value = row[i-1].value
	r += 1

wb.save("tweets_similar_to_retweeted.xlsx")
'''

def compare_dates(first,second):
	if "2016-" in first:
		if "2016-" in second:
			return first < second
		else:
			return True
	else:
		if "2016-" in second:
			return False
		else:
			for day in days:
				first = first.strip(day)
				second = second.strip(day)
			return first<second


ortct= 0
odate = 0
crtct = 0
cdate = 0
after = True
r = 1
for row in ws.rows:
	if row[0].value == "ORIGINAL TWEET":
		if not after:
			first_tweet_count += 1
		total += 1
		after = False
		odate = row[4].value
		ortct = row[3].value
	if type(row[0].value) is not type(None) and not after:
		crtct = row[3].value
		cdate = row[4].value
		if crtct < ortct:
			after = not compare_dates(odate,cdate)
		if after:
			print r
	r += 1

if not after:
	first_tweet_count += 1

print first_tweet_count/total

