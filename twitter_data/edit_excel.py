import openpyxl

wb = openpyxl.load_workbook("tweets_similar_to_retweeted.xlsx")
old_ws = wb.active
max_row = old_ws.get_highest_row()
max_col = old_ws.get_highest_column()
wb.create_sheet("new_sheet",index=0)
new_sheet = wb.get_sheet_by_name("new_sheet")

r = 1
for row in old_ws.rows:
	if row[0].value=="ORIGINAL TWEET":
		new_sheet.cell(row = r, column = 1).value = ''
		r += 1
	for i in range(1,max_col+1):
		new_sheet.cell(row=r,column=i).value = row[i-1].value
	r += 1

wb.save("tweets_similar_to_retweeted.xlsx")


